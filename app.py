import streamlit as st
import pandas as pd
import google.generativeai as genai
import json
import fitz
import PIL.Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# ── Setup ──
GEMINI_KEY = os.environ.get('GEMINI_KEY')
genai.configure(api_key=GEMINI_KEY)
model = genai.GenerativeModel("gemini-2.5-flash")

def extract_from_image(img):
    prompt = """אתה מומחה חשבונאות ישראלי. נתח את המסמך והחזר JSON בלבד.

חשוב: אם זו קבלה שהוצאת ללקוח (אתה המוכר) — document_type = "income"
אם זו חשבונית שקיבלת מספק (אתה הקונה) — document_type = "expense"

{
    "document_type": "expense או income",
    "supplier": "שם הספק או הלקוח",
    "amount": 0.0,
    "vat": 0.0,
    "date": "DD/MM/YYYY",
    "category": "ציוד_משרדי / תוכנה / נסיעות / אירוח / ייעוץ / שיווק / אחר",
    "tax_recognized_percent": 100,
    "confidence": "high/medium/low",
    "notes": ""
}

חוקי מס הכנסה ישראל:
- אירוח ומסעדות: מוכר 50%
- רכב ודלק: מוכר 50%
- ספקים זרים: ציין Reverse Charge בהערות
- כל השאר: מוכר 100%

החזר JSON בלבד ללא טקסט נוסף."""
    response = model.generate_content([prompt, img])
    text = response.text.strip()
    if "```" in text:
        text = text.split("```")[1]
        if text.startswith("json"):
            text = text[4:]
    return json.loads(text.strip())

def pdf_to_image(file_bytes):
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    page = doc[0]
    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
    return PIL.Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

def create_excel(expenses, income, flagged):
    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    flag_fill = PatternFill("solid", fgColor="FFE0E0")

    ws1 = wb.active
    ws1.title = "הוצאות"
    headers = ["קובץ","ספק","תאריך","קטגוריה","סכום","מע\"מ","מוכר %","סכום מוכר","הערות"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total_amount = 0
    total_recognized = 0
    for row, r in enumerate(expenses, 2):
        pct = r.get("tax_recognized_percent", 100)
        amt = float(r.get("amount", 0))
        rec = amt * pct / 100
        total_amount += amt
        total_recognized += rec
        vals = [r.get("source_file",""), r.get("supplier",""), r.get("date",""),
                r.get("category",""), amt, float(r.get("vat",0)), f"{pct}%", rec, r.get("notes","")]
        for col, v in enumerate(vals, 1):
            ws1.cell(row=row, column=col, value=v)
        if r.get("confidence") == "low":
            for col in range(1, len(headers)+1):
                ws1.cell(row=row, column=col).fill = flag_fill

    sum_row = len(expenses) + 2
    ws1.cell(row=sum_row, column=4, value="סה\"כ").font = Font(bold=True)
    ws1.cell(row=sum_row, column=5, value=total_amount).font = Font(bold=True)
    ws1.cell(row=sum_row, column=8, value=total_recognized).font = Font(bold=True)

    for col in range(1, 10):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    if income:
        ws2 = wb.create_sheet("הכנסות")
        inc_headers = ["קובץ","לקוח","תאריך","סכום","מע\"מ","הערות"]
        for col, h in enumerate(inc_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
        for row, r in enumerate(income, 2):
            ws2.cell(row=row, column=1, value=r.get("source_file",""))
            ws2.cell(row=row, column=2, value=r.get("supplier",""))
            ws2.cell(row=row, column=3, value=r.get("date",""))
            ws2.cell(row=row, column=4, value=float(r.get("amount",0)))
            ws2.cell(row=row, column=5, value=float(r.get("vat",0)))
            ws2.cell(row=row, column=6, value=r.get("notes",""))

    if flagged:
        ws3 = wb.create_sheet("לבדיקה ידנית")
        ws3.cell(row=1, column=1, value="קבצים הדורשים בדיקה:").font = Font(bold=True, color="FF0000")
        for row, r in enumerate(flagged, 2):
            ws3.cell(row=row, column=1, value=r.get("source_file",""))
            ws3.cell(row=row, column=2, value=r.get("notes",""))

    excel_path = "/tmp/דוח_הוצאות.xlsx"
    wb.save(excel_path)
    return excel_path

# ── UI ──
st.set_page_config(page_title="מחולל דוח מס הכנסה", page_icon="📄", layout="wide")
st.title("📄 מחולל דוח הוצאות למס הכנסה")
st.markdown("העלה קבלות וחשבוניות — המערכת תסרוק, תסווג ותפיק דוח Excel מוכן להגשה")

col_left, col_right = st.columns(2)

with col_left:
    st.subheader("📤 חשבוניות הוצאה")
    st.caption("חשבוניות שקיבלת מספקים — תשלומים שלך")
    invoice_files = st.file_uploader(
        "העלה חשבוניות",
        accept_multiple_files=True,
        type=["pdf", "jpg", "jpeg", "png"],
        key="invoices"
    )

with col_right:
    st.subheader("📥 קבלות הכנסה")
    st.caption("קבלות שהוצאת ללקוחות — תשלומים שקיבלת")
    receipt_files = st.file_uploader(
        "העלה קבלות",
        accept_multiple_files=True,
        type=["pdf", "jpg", "jpeg", "png"],
        key="receipts"
    )

all_files = []
if invoice_files:
    all_files += [(f, "expense") for f in invoice_files]
if receipt_files:
    all_files += [(f, "income") for f in receipt_files]

st.write(f"DEBUG: {len(all_files)} קבצים מוכנים")

button_clicked = st.button("🚀 עבד והפק דוח", type="primary", disabled=len(all_files) == 0)

if button_clicked and all_files:
    results = []
    errors = []
    progress = st.progress(0)
    status = st.empty()

    for i, (f, hint) in enumerate(all_files):
        status.text(f"🔄 מעבד ({i+1}/{len(all_files)}): {f.name}")
        progress.progress((i+1) / len(all_files))
        try:
            if f.name.lower().endswith(".pdf"):
                img = pdf_to_image(f.read())
            else:
                img = PIL.Image.open(f)
            data = extract_from_image(img)
            data["source_file"] = f.name
            results.append(data)
            st.write(f"✅ {f.name}: {data.get('document_type')} — {data.get('supplier')}")
        except Exception as e:
            errors.append({"file": f.name, "error": str(e)})
            st.write(f"❌ {f.name}: {str(e)}")

    progress.empty()
    status.empty()

    if results:
        expenses = [r for r in results if r.get("document_type") == "expense"]
        income   = [r for r in results if r.get("document_type") == "income"]
        flagged  = [r for r in results if r.get("confidence") == "low"]

        st.markdown("---")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("📁 סה\"כ", len(results))
        c2.metric("📤 הוצאות", len(expenses))
        c3.metric("📥 הכנסות", len(income))
        c4.metric("⚠️ לבדיקה", len(flagged))

        if expenses:
            st.subheader("📤 הוצאות")
            df_exp = pd.DataFrame([{
                "קובץ": r.get("source_file",""),
                "ספק": r.get("supplier",""),
                "תאריך": r.get("date",""),
                "קטגוריה": r.get("category",""),
                "סכום": r.get("amount",0),
                "מע\"מ": r.get("vat",0),
                "מוכר %": f"{r.get('tax_recognized_percent',100)}%",
                "סכום מוכר": r.get("amount",0) * r.get("tax_recognized_percent",100) / 100
            } for r in expenses])
            st.dataframe(df_exp, use_container_width=True)
            total = sum(r.get("amount",0) for r in expenses)
            recognized = sum(r.get("amount",0) * r.get("tax_recognized_percent",100) / 100 for r in expenses)
            st.info(f"**סה\"כ הוצאות:** {total:.2f}  |  **סה\"כ מוכר למס:** {recognized:.2f}")

        if income:
            st.subheader("📥 הכנסות")
            df_inc = pd.DataFrame([{
                "קובץ": r.get("source_file",""),
                "לקוח": r.get("supplier",""),
                "תאריך": r.get("date",""),
                "סכום": r.get("amount",0),
                "מע\"מ": r.get("vat",0),
                "הערות": r.get("notes","")
            } for r in income])
            st.dataframe(df_inc, use_container_width=True)

        if errors:
            st.warning(f"⚠️ {len(errors)} קבצים לא עובדו:")
            for e in errors:
                st.write(f"• {e['file']}: {e['error']}")

        if flagged:
            st.warning("⚠️ קבצים לבדיקה ידנית:")
            for r in flagged:
                st.write(f"• {r.get('source_file','')} — {r.get('notes','')}")

        st.markdown("---")
        excel_path = create_excel(expenses, income, flagged)
        with open(excel_path, "rb") as xf:
            st.download_button(
                "⬇️ הורד דוח Excel",
                data=xf.read(),
                file_name="דוח_הוצאות.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        st.success("✅ הדוח מוכן להורדה!")
